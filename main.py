#! /usr/bin/env python
# -*- coding: utf-8 -*-

import string
from openpyxl import load_workbook

"""
openpyxl 不支持old .xls file format
how to convert .xls file to .xlsx?
"""
excel_file = 'wacai_2019-05-2019-05.xlsx'
sheet_names = ['支出', '收入', '转账', '借入借出', '收款还款']



class DealExcel:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.data = load_workbook(self.excel_path)
        self.sheetnames = self.data.sheetnames

    def convert_alphabet_to_int(self, alphabet):
        return string.ascii_uppercase.index(alphabet) + 1

    def convert_int_to_alphabet(self, num):
        return string.ascii_uppercase[num - 1]

    def read_by_sheetname(self, sheetname):
        sheetname_list = []
        if sheetname in self.sheetnames:
            wb = self.data[sheetname]
            for j in range(1, wb.max_row + 1):
                line_list = []
                for i in range(1, wb.max_column + 1):
                    letter = self.convert_int_to_alphabet(i)
                    cell_value = wb[letter+str(j)].value
                    line_list.append(cell_value)
                sheetname_list.append(line_list)
        return sheetname_list

    def read_all_sheetnames(self):
        data_dict = {}
        for sheetname in self.sheetnames:
            sheet_data = self.read_by_sheetname(sheetname)
            data_dict[sheetname] = sheet_data
        return  data_dict

if __name__ == "__main__":
    deal_excel = DealExcel(excel_file)
    #deal_excel.read_by_sheetname('收入')
    content = deal_excel.read_all_sheetnames()
    print(content)
